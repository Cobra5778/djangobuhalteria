#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.shortcuts import render
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from .remote import TTK_all, nav_mounth_scroll, TTK_summ, SQL_to_CSV, RTK_summ, bill_groups
import os, csv
import mimetypes
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
from django.contrib.auth import authenticate, login, logout

#MY_PATH = "/home/earth/PycharmProjects/djangobuhalteria/buhal/"
MY_PATH = os.path.abspath(os.curdir) + '/buhal/'
#shablon = 'Ballance.xlsx'
SHABLON_BALLANCE = 'Ballance.xlsx'
SHABLON_RTK_CUS = 'RTK_CUS.xlsx'
SHABLON_RTK_BIL = 'RTK_BIL.xlsx'
SHABLON_AKT_TTK= 'akt_ttk.xlsm'

def wellcome(request):
    error_message = ""
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
        return render(request, 'index.html', {'error_message': error_message, })
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))

def logins(request):
    error_message = ""
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                # Redirect to a success page.
                return HttpResponseRedirect(reverse('wellcome', args=[]))
                # return HttpResponse( Messages)
            else:
                # Return a 'disabled account' error message
                error_message = "Ваш аккаунт отключен, Дозвиданья!"
                return render(request, 'login.html', {'error_message': error_message, })
        else:
            # Return an 'invalid login' error message.
            error_message = "Не верный логин или пароль"
            return render(request, 'login.html', {'error_message': error_message, })
    else:
        return render(request, 'login.html', {'error_message': error_message, })


def logouts(request):
    logout(request)
    # Redirect to a success page.
    return HttpResponseRedirect(reverse('logins', args=[]))

def nav_mounth(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    table_prefix_scroll = nav_mounth_scroll(range_data)
    return render(request, 'query/nav_month.html', {'table_prefix_scroll': table_prefix_scroll, })


def nav_ballansy(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    table_prefix_scroll = nav_mounth_scroll(range_data)
    return render(request, 'query/nav_ballansy.html', {'table_prefix_scroll': table_prefix_scroll, })


def nav_mounth_TTK(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    table_prefix_scroll = nav_mounth_scroll(range_data)
    return render(request, 'query/nav_month_TTK.html', {'table_prefix_scroll': table_prefix_scroll, })


def wellcome(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    return render(request, 'main.html', {'error_message': '', 'full_name': full_name, })


def ballansy(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    return render(request, 'ballasy.html', {'error_message': '', 'full_name': full_name, 'table_data': ''})


def TTK_main(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    return render(request, 'ttk_main_from.html', {'error_message': '', 'full_name': full_name, 'table_data': ''})


def RTK_main(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    return render(request, 'rtk_main_from.html', {'error_message': '', 'full_name': full_name, 'table_data': ''})


def compare_sf(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    table_data = bill_groups();
    return render(request, 'compare_sf.html', {'error_message': '', 'full_name': full_name, 'table_data': table_data})


def TTKakt(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    table_data = TTK_all(range_data)
    summ = 0
    for item in table_data:
        summ += item[11]
    return render(request, 'query/TTK-akt.html', {'table_data': table_data, 'summ': summ, })


def ballansy_uriki(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    SQL = """select C.date1, C.date2, C.title, CP.val,
                  concat(CB.mm, '.', CB.yy) as period,
            case 
            when
            ((CB.yy*100 + CB.mm)>={})
            then concat(CB.summa1) 
            else 
            concat(CB.summa1+CB.summa2-CB.summa3-CB.summa4) end as Ostatok, C.comment
            from contract C left join contract_balance CB on CB.cid=C.id and
                CB.yy=(select max(yy) from contract_balance where cid=C.id)    and 
                CB.mm=(select max(mm) from contract_balance where cid=C.id and yy=CB.yy) , contract_parameter_type_1 CP
            where  C.id=CP.cid and (CP.pid=10 or CP.pid=33) and CB.summa1 is not NULL AND C.fc = 1
            limit 5000""".format(range_data)
    # print (SQL)
    table_data = SQL_to_CSV(SQL)
    return render(request, 'query/ballansy-yuiriki.html', {'table_data': table_data, })


def sf_result(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    summSF = 0.0
    summBILL = 0.0
    addOnSQL = ""
    flagFirst = True
    range_data = request.GET['range'] #.encode('UTF-8')
    groups = request.GET['res'] #.encode('UTF-8')
    subs = request.GET['subs'] #.encode('UTF-8')
    print
    subs
    if subs == "true":
        addOn2SQL = " AND (C.scid = -1 or C.scid = 0)"
    else:
        addOn2SQL = ""
    if groups != "":
        addOnSQL = " AND ("
        for group in groups:
            if flagFirst:
                addOnSQL += "C.gr&(1<<{})".format(ord(group))  # -- and (C.gr&(1<<1) or C.gr&(1<<7))
                flagFirst = False
            else:
                addOnSQL += " OR C.gr&(1<<{})".format(ord(group))
        addOnSQL += ")"
    # print (addOnSQL)
    SQL = """SELECT C.gr, C.title, C.comment, CB.cid, CB.yy, CB.mm, BI.format_number, BI.summ as summaSF, CB.summa3 as summaBilling, (BI.summ-CB.summa3) as razn, C.scid
                from contract_balance CB
                    left join bill_invoice_data_14 BI on (CB.cid=BI.cid and CB.yy=BI.yy and CB.mm=BI.mm+1)
                    left join contract C on (C.id=CB.cid)
                where CB.yy={1} and CB.mm={2} and (BI.summ<>CB.summa3 or (BI.summ is NULL and CB.summa3>0)) {4} {3}
            limit 5000""".format(range_data, range_data[:4], range_data[4:], addOnSQL, addOn2SQL)
    # print (SQL)
    table_data = SQL_to_CSV(SQL)
    for in_Date in table_data:
        try:
            summSF += float(str(in_Date[7]))
        except:
            summSF += 0
        try:
            summBILL += float(str(in_Date[8]))
        except:
            summBILL += 0
    return render(request, 'query/sf_result.html',
                  {'table_data': table_data, 'summSF': summSF, 'summBILL': summBILL, 'summ': (summSF - summBILL)})


def ballansy_fiziki(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    SQL = """select C.date1, C.date2, C.title, CP.val,
                  concat(CB.mm, '.', CB.yy) as period,
            case 
            when
            ((CB.yy*100 + CB.mm)>={})
            then concat(CB.summa1) 
            else 
            concat(CB.summa1+CB.summa2-CB.summa3-CB.summa4) end as Ostatok, C.comment
            from contract C left join contract_balance CB on CB.cid=C.id and
                CB.yy=(select max(yy) from contract_balance where cid=C.id)    and 
                CB.mm=(select max(mm) from contract_balance where cid=C.id and yy=CB.yy) , contract_parameter_type_1 CP
            where  C.id=CP.cid and (CP.pid=10 or CP.pid=33) and CB.summa1 is not NULL AND C.fc = 0
            limit 5000""".format(range_data)
    # print (SQL)
    table_data = SQL_to_CSV(SQL)
    return render(request, 'query/ballansy-fiziki.html', {'table_data': table_data, })


def RTKsumm(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    SQL = """SELECT CASE
                WHEN sid=12 THEN "MГ"
                WHEN sid=14 THEN "MH"
                WHEN sid=80 THEN "B3"
            END as Type,
            sum(LS.session_time)/60 as minutes, sum(LS.oper_session_cost) as cost 
            FROM log_session_8_{} LS, contract_tariff CT
            WHERE LS.sid <> 2
                AND LS.cid = CT.cid
                AND CT.tpid=145 /* Код тарифа  TTK-95,88 Ростелеком-137 Ростелеком-145*/
            GROUP BY sid """.format(range_data)
    # print (SQL)
    table_data = SQL_to_CSV(SQL)
    summ = 0
    for item in table_data:
        summ += item[2]
    return render(request, 'query/RTK_sum_table.html', {'table_data': table_data, 'summ': summ, })


def downloadCSV_TTK(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    my_File_name = "my_filename.csv"
    range_data = request.GET['range']#.encode('UTF-8')
    ANI = """SELECT
                c.`title` as `Num_Dog`,    /*Номер договора ТрансТелекома с пользователем*/
                "-" as `Num_Sub`,                     /*№ Дог. оператора с Субагентом*/
                "472" as `ABC`,                    /*ABC код*/
                login_alias as `AOH`,             /*АОН*/
                c.`date1` as `Date_Podkl`,                 /*Дата подключения*/
                c.`date2` as `Date_Otkl`                 /*Дата отключения*/
            FROM `user_login_8` ul, `user_alias_8` ua, `contract` c
            WHERE ua.`login_id` = ul.`id`
                AND ul.`cid` = c.`id` AND c.title like "DA000287%" 
            ORDER BY c.`title`
            limit 100000"""
    BIL = """SELECT  
                "DA000287" as `NumDogTrans`,                                /*Нормер дого. с ТрансТелекомом*/
                "-" as `N_DogSup`,                                         /*№ Дог. оператора с Субагентом*/
                c.`title` as `NumDog`,                            /*Номер договора ТрансТелекома с пользователем*/
                bl.`format_number` as `N_SF`,                                 /*Номер Счет-фактуры*/
                bl.`format_number` as `N_AKT`,                                /*Номер Акта*/
                bl.`create_dt` as `date_doc`,                                  /*Дата выставления*/
                DATE_ADD(bl.`create_dt`, INTERVAL 30 DAY) as `srok_opl`,                     /*Срок оплаты*/
                 c.`title` as `KodPol`,                         /*Код пользователя*/
                643 as `Valuta`,                                         /*Код валюты счета*/
                18 as `NDS`,                                         /*Ставка НДС*/
                IF(ls.`zone`=2,"12","10") as `Kod_Usl`,                                        /*Код услуги*/
                concat(bl.`yy`,"-", bl.`mm` + 1, "-01") as `Date_Usl`,                     /*Дата услуги*/
                round(SUM(ls.`session_cost`)/1.2,2) as `Cost`,                 /* Цена без НДС */
                ROUND(SUM(ls.`round_session_time`)/60) as `kol_ed`,                     /*Кол. минут*/
                "14440" as `Mesto`,                                        /*Место оказания услуг*/
                ROUND(SUM(ls.`session_cost`),2)  as `AllCost`                             /*Сумма с НДС*/
                FROM  `bill_invoice_data_14` bl, `contract` c, `log_session_8_{}` ls  /*МЕСЯЦ*/
            WHERE bl.`type` = 11 /* 21 Ростелеком; 11 - ТТК */
                AND c.`id` = bl.`cid`
                AND c.`id` = ls.`cid`
                AND (ls.`zone`=4 OR ls.`zone`=2)
                AND ls.`session_cost` > 0
                AND bl.`yy`={}
                AND bl.`mm`={} /* МЕСЯЦ - 1 */
            GROUP BY NumDogTrans, N_DogSup, NumDog, N_SF, N_AKT, date_doc, srok_opl, NumDog, Valuta, NDS, Kod_Usl, Date_Usl, Mesto
            limit 100000""".format(range_data, range_data[:4], int(range_data[4:]) - 1)
    CUS = """SELECT 
                    c.`title` as Nomer_Dog,            /*Номер договора в базе*/
                    1 as Type_Podluch,                        /*Тип договора 1-абонент*/
                    concat_ws('',NAIMENOVANIE.`val`,FIO.`val`) as NAIMENOVANIE,    /*Наименование (или ФИО)*/
                    concat_ws('',UR_ADRES.`val`,PROPISKA.`val`) as ADRESS,        /*Юридический адрес или Прописка*/
                    INN.`val` as INN,                        /*ИНН для юриков*/
                    KPP.`val` as KPP,                        /*КПП для юриков*/
                    2 as DIPLOMAT,                            /*Дипломат 1-да 2-нет */
                    (fc + 1) as UR_STATUS,                        /* Юридический статус 1-физик 2-юрик */
                    0 as RESIDENT,                            /* Резидент 0-нет 1-да */
                    643 as NAC_PRIN,                        /* Национальная принадлежность */
                    c.`date1` as Date_Begin,                        /* Дата заключения договора */
                    c.`date2` as Date_End,                        /* Дата прекращения договора */
                    1 as DOG_Type,                            /* Тип договора КЛИЕНТСКИЙ*/
                    '14 440' as Mesto_Zakluch,                    /* Место заключения договора */
                    0 as SOGLASIE                            /* Согласие на лич.данные 0-нет, 1-да */
                FROM  `contract_module` CM, `contract` c
                    LEFT JOIN  `contract_parameter_type_1` NAIMENOVANIE ON (c.`id` = NAIMENOVANIE.`cid` AND NAIMENOVANIE.`pid` = 10 )
                    LEFT JOIN  `contract_parameter_type_1` FIO ON (c.`id` = FIO.`cid` AND FIO.`pid` = 33 )
                    LEFT JOIN  `contract_parameter_type_1` UR_ADRES ON (c.`id` = UR_ADRES.`cid` AND UR_ADRES.`pid` = 17 )
                    LEFT JOIN  `contract_parameter_type_1` PROPISKA ON (c.`id` = PROPISKA.`cid` AND PROPISKA.`pid` = 35 )
                    LEFT JOIN  `contract_parameter_type_1` INN ON (c.`id` = INN.`cid` AND INN.`pid` = 25 )
                    LEFT JOIN  `contract_parameter_type_1` KPP ON (c.`id` = KPP.`cid` AND KPP.`pid` = 32 )
                WHERE CM.`mid` = 8 AND c.`id` = CM.`cid` and c.title LIKE "DA000287%"
                limit 100000 """
    if request.GET['type'] == "ANI": #.encode('UTF-8')
        mySQL = ANI
        out_filename = "DA000287_ANI_{}_{}.csv".format(range_data[:4], range_data[4:])
    if request.GET['type'] == "CUS": #.encode('UTF-8')
        mySQL = CUS
        out_filename = "DA000287_CUS_{}_{}.csv".format(range_data[:4], range_data[4:])
    if request.GET['type'] == "BIL": #.encode('UTF-8')
        mySQL = BIL
        out_filename = "DA000287_BIL_{}_{}.csv".format(range_data[:4], range_data[4:])
    csv_file = open(my_File_name, 'w')
    writer = csv.writer(csv_file, delimiter=str(u';'))
    writer.writerows(SQL_to_CSV(mySQL))
    csv_file.close()
    csv_file = open(my_File_name, "rb")
    response = HttpResponse(csv_file)
    file_type = mimetypes.guess_type(my_File_name)
    if file_type is None:
        file_type = 'application/octet-stream'
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(my_File_name).st_size)
    response['Content-Disposition'] = "attachment; filename={}".format(out_filename)
    csv_file.close()
    os.remove(my_File_name)
    return response;


def my_downloadfile(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    #MY_PATH = "/home/cobra/workspace/manage-radio/buhalteriya/"
    #MY_PATH = "~/PycharmProjects/djangobuhalteria/buhal"
    # excel_file_name = question_id
    excel_file_name = MY_PATH + "test.xls"
    fp = open(excel_file_name, "rb")
    response = HttpResponse(fp.read())
    fp.close()
    file_type = mimetypes.guess_type(excel_file_name)
    if file_type is None:
        file_type = 'application/octet-stream'
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(excel_file_name).st_size)
    response['Content-Disposition'] = "attachment; filename=report.csv"
    # os.remove(excel_file_name
    return response;
    # return render(request, 'main.html', { 'error_message': '', 'full_name': excel_file_name, })


def Summ_Stoim_TTK(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    response = HttpResponse()
    response.content = TTK_summ(range_data)
    return response


def Summ_Stoim_RTK(request):
    if request.user.is_authenticated:
        # Do something for authenticated users.
        full_name = request.user.get_full_name()
    else:
        # Do something for anonymous users.
        return HttpResponseRedirect(reverse('logins', args=[]))
    range_data = request.GET['range']#.encode('UTF-8')
    response = HttpResponse()
    response.content = RTK_summ(range_data)
    return response


def export_movies_to_xlsx_old(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response = HttpResponse()
    response['Content-Disposition'] = 'attachment; filename=test-movies.xlsx'
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Movies'
    # Define the titles for columns
    columns = ['ID', 'Title', 'Description', 'Length', 'Rating', 'Price', ]
    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # workbook.save(response)
    response.content = save_virtual_workbook(workbook)
    return response


def export_TTK_to_xlsm(request):
    CHET = ""
    SF = ""
    AKT = ""
    file_name = SHABLON_AKT_TTK
    #MY_PATH = "/home/cobra/workspace/manage-radio/buhalteriya/"
    range_data = request.GET['range']#.encode('UTF-8')
    CHET = request.GET['num_chet']#.encode('UTF-8')
    SF = request.GET['num_sf']#.encode('UTF-8')
    AKT = request.GET['num_akt']#.encode('UTF-8')
    # workbook = load_workbook(MY_PATH + 'test2.xlsx')
    workbook = load_workbook(MY_PATH + file_name, keep_vba=True)
    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response = HttpResponse()
    response['Content-Disposition'] = 'attachment; filename={}_TTK_akt.xlsm'.format(range_data)
    # workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.worksheets[0]
    # worksheet = workbook.active
    # worksheet.title = 'Movies'
    # Define the titles for columns
    table_data = TTK_all(range_data)
    row_num = 1
    for columns in table_data:
        row_num += 1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
    # workbook.active = 1
    # worksheet = workbook.active
    # worksheet = workbook.worksheets[1]
    # worksheet['B{}'.format(22)] = '=Данные!R[-19]C[-1]'
    # worksheet.insert_rows(9,row_num)
    # for i in range(9, row_num + 9):
    #    worksheet['A{}'.format(i)] = '==Данные!A{}'.format(i-7)
    #    worksheet['B{}'.format(i)] = '=="№"&Данные!B{0}&" от "&Данные!C{0}'.format(i-7)
    #    cell = worksheet.cell(row=i, column=1).value = "=СУММ(A7; B7)"
    # cell = worksheet.cell(row=i, column=2).value = """=СЦЕПИТЬ("№";Данные!R[-7]C;" от "; ТЕКСТ(Данные!R[-7]C[1];"ДД.ММ.ГГГГ\г\."))"""
    # worksheet.delete_rows(9,10)
    # for row in table_data:
    #    worksheet.append(row)
    # workbook.save(response)
    worksheet['S3'] = '{}'.format(CHET)  # Счет
    worksheet['S4'] = '{}'.format(SF)  # СФ
    worksheet['S5'] = '{}'.format(AKT)  # АКТ
    response.content = save_virtual_workbook(workbook)
    return response


def export_ballanse_xlsx(request):
    #MY_PATH = "/home/cobra/workspace/manage-radio/buhalteriya/"
    range_data = request.GET['range']#.encode('UTF-8')
    mySQL = """SELECT C.date1, C.date2, C.title, CP.val , concat(CB.summa1, 'руб.') as in_Ostatok, 
          concat(CB.summa2, 'руб.') as Prihod,  concat(CB.summa3, 'руб.') as Narabotka,
          concat(CB.summa4, 'руб.') as Pashod, concat(CB.summa1+CB.summa2-CB.summa3-CB.summa4, 'руб.') as out_ostatok,
          concat('на ', CB.mm, '.', CB.yy) as period,  C.comment,
            case
                when
                ((CB.yy*100 + CB.mm)>={})
                then concat(CB.summa1)
                else
                concat(CB.summa1+CB.summa2-CB.summa3-CB.summa4) end as Ostatok
        from contract C left join contract_balance CB on CB.cid=C.id and
                    CB.yy=(select max(yy) from contract_balance where cid=C.id) and 
                    CB.mm=(select max(mm) from contract_balance where cid=C.id and yy=CB.yy) , contract_parameter_type_1 CP
        where  C.id=CP.cid and (CP.pid=10 or CP.pid=33) and CB.summa1 is not NULL AND C.fc = {}
        limit 7000"""
    out_filename = "Ballance_{}_{}.xlsx".format(range_data[:4], range_data[4:])
    shablon = SHABLON_BALLANCE
    workbook = load_workbook(MY_PATH + shablon)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response['Content-Disposition'] = 'attachment; filename={}'.format(out_filename)  # 0301_25_1176_17_CUS_2019_10.xls

    worksheet = workbook.worksheets[0]
    table_data = SQL_to_CSV(mySQL.format(range_data, 0), numeric=False)  # Сюда вставить запрос
    row_num = 0
    for columns in table_data:
        row_num += 1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

    worksheet = workbook.worksheets[1]
    table_data = SQL_to_CSV(mySQL.format(range_data, 1), numeric=False)  # Сюда вставить запрос
    row_num = 0
    for columns in table_data:
        row_num += 1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

    response.content = save_virtual_workbook(workbook)
    return response


def export_RTK_to_xlsx(request):
    #MY_PATH = "/home/cobra/workspace/manage-radio/buhalteriya/"
    range_data = request.GET['range']#.encode('UTF-8')
    out_filename = ""
    mySQL = ""
    shablon = ""

    BIL = """ SELECT 
            "0301/25/1176-17" as `NumDogTrans`,                /*Нормер дого. с ТрансТелекомом*/
            c.`title` as `NumDog`,                            /*Номер договора ТрансТелекома с пользователем*/
            bl.`format_number` as `N_SF`,                     /*Номер Счет-фактуры*/
            bl.`format_number` as `N_AKT`,                    /*Номер Акта*/
            DATE_FORMAT(bl.`create_dt`,'%d-%m-%Y') as `date_doc`,  /*Дата выставления*/
            2 as `NDS`,                                         /*Ставка НДС*/
            CASE
                WHEN ls.sid=12 THEN 12 /* Код услуги: 10 - международная связь, 12 - междугородная связь, 14 - внутризоновая связь */
                WHEN ls.sid=14 THEN 10
                WHEN ls.sid=80 THEN 14
                ELSE 0
            END as `Kod_Usl`, /* Код услуги: 10 - международная связь, 12 - междугородная связь, 14 - внутризоновая связь */ 
            DATE_FORMAT(concat(bl.`yy`,"-", bl.`mm` + 1, "-01"),'%d-%m-%Y') as `Date_Usl`, /*Дата услуги*/
            round(SUM(ls.`session_cost`),4) as `Cost`,         /* Цена без НДС */
            ROUND(SUM(ls.`round_session_time`)/60) as `kol_ed`,     /*Кол. минут*/
            "14440" as `Mesto`,                    /*Место оказания услуг*/
            1  as `Vkl_NDS`,
            IF(c.fc=0,"26","01") as UR_STATUS     /*Код вида операции*/
        FROM  `bill_invoice_data_14` bl
            LEFT JOIN `log_session_8_{}` ls  /*МЕСЯЦ*/ ON bl.cid=ls.cid AND ls.sid <> 2 AND ls.`session_cost` > 0,
         `contract` c
        WHERE bl.`type` = 23 /* 23, 21 Ростелеком; 11 - ТТК */
            AND c.`id` = bl.`cid` 
            AND bl.`yy`={}
            AND bl.`mm`={} /* МЕСЯЦ - 1 */
        GROUP BY N_SF, ls.sid
        limit 100000 """.format(range_data, range_data[:4], int(range_data[4:]) - 1)

    CUS = """SELECT 
                c.`title` as Nomer_Dog,            /*Номер договора Ростелеком с Абонентом */
                concat_ws('',NAIMENOVANIE.`val`,FIO.`val`) as NAIMENOVANIE,    /* Полное наименование Абонента (или ФИО) */
                concat_ws('',UR_ADRES.`val`,PROPISKA.`val`) as ADRESS,        /* Юридический адрес или прописка */
                INN.`val` as INN,                        /*ИНН для юриков*/
                KPP.`val` as KPP,                        /*КПП для юриков*/
                (fc + 1) as UR_STATUS,                        /* Юридический статус 1-физик 2-юрик */
                DATE_FORMAT(c.`date1`,'%d-%m-%Y') as Date_Begin,                        /* Дата заключения договора */
                DATE_FORMAT(c.`date2`,'%d-%m-%Y') as Date_End,                        /* Дата прекращения договора */
                '14 440' as Mesto_Zakluch,                    /* Место заключения договора */
                0 as SOGLASIE,                            /* Согласие на лич.данные 0-нет, 1-да */
                0 as RESIDENT,                            /* Резидент 0-нет 1-да */
                'N' as KORREKTIROVKA                            /*Признак корректировки данных о сущ. Пользователе Y- корректировка*/
            FROM  `contract_module` CM, contract_tariff CT,`contract` c
                LEFT JOIN  `contract_parameter_type_1` NAIMENOVANIE ON (c.`id` = NAIMENOVANIE.`cid` AND NAIMENOVANIE.`pid` = 10 )
                LEFT JOIN  `contract_parameter_type_1` FIO ON (c.`id` = FIO.`cid` AND FIO.`pid` = 33 )
                LEFT JOIN  `contract_parameter_type_1` UR_ADRES ON (c.`id` = UR_ADRES.`cid` AND UR_ADRES.`pid` = 17 )
                LEFT JOIN  `contract_parameter_type_1` PROPISKA ON (c.`id` = PROPISKA.`cid` AND PROPISKA.`pid` = 35 )
                LEFT JOIN  `contract_parameter_type_1` INN ON (c.`id` = INN.`cid` AND INN.`pid` = 25 )
                LEFT JOIN  `contract_parameter_type_1` KPP ON (c.`id` = KPP.`cid` AND KPP.`pid` = 32 )
            WHERE c.id = CT.cid 
                AND CT.tpid=145 /* Код тарифа  TTK-95,88 Ростелеком-137 Ростелеком-145*/     
                AND CM.`mid` = 8 AND c.`id` = CM.`cid`
            limit 100000"""

    if request.GET['type'] == "CUS": #.encode('UTF-8')
        mySQL = CUS
        out_filename = "0301_25_1176_17_CUS_{}_{}.xlsx".format(range_data[:4], range_data[4:])
        shablon = SHABLON_RTK_CUS
    if request.GET['type'] == "BIL": #.encode('UTF-8')
        mySQL = BIL
        out_filename = "0301_25_1176_17_BIL_{}_{}.xlsx".format(range_data[:4], range_data[4:])
        shablon = SHABLON_RTK_BIL

    workbook = load_workbook(MY_PATH + shablon)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    # response = HttpResponse()
    response['Content-Disposition'] = 'attachment; filename={}'.format(out_filename)  # 0301_25_1176_17_CUS_2019_10.xls
    worksheet = workbook.worksheets[0]
    table_data = SQL_to_CSV(mySQL, numeric=True)  # Сюда вставить запрос
    row_num = 0
    for columns in table_data:
        row_num += 1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
    response.content = save_virtual_workbook(workbook)
    return response